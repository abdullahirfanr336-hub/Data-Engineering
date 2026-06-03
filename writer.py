import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class ExcelWriter:
    """
    Generates the final Excel file with styling and conditional formatting.
    """
    def __init__(self, output_path):
        self.output_path = output_path
        self.red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    def save(self, df):
        """
        Saves the DataFrame to Excel and applies a RED highlight to empty/null cells.
        """
        # Save using pandas first for the structure
        df.to_excel(self.output_path, index=False)

        # Re-open with openpyxl to apply conditional formatting (Rule #4)
        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)
        ws = wb.active

        # Identify required columns or just check all for NULL/EMPTY
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    cell.fill = self.red_fill

        wb.save(self.output_path)
        print(f"File saved to {self.output_path} with formatting.")

if __name__ == "__main__":
    print("ExcelWriter initialized.")
